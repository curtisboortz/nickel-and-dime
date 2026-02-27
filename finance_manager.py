"""
Finance Manager - Portfolio tracking with live prices, allocation metrics, and budget.
Run this script to fetch prices, update the Excel workbook, and generate the dashboard.

  python finance_manager.py           # One-time update
  python finance_manager.py --watch   # Watch mode: stocks/crypto every 5 min (stocks during market hours),
                                      # metals 4x/day. Ctrl+C to stop.
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# Load .env so API keys can live in env instead of config.json
try:
    from dotenv import load_dotenv
    _base = Path(__file__).resolve().parent
    load_dotenv(_base / ".env")
except ImportError:
    pass

# Metals refresh: 4x/day = every 6 hours (stays under GoldAPI 100 req/month)
METALS_INTERVAL_HOURS = 6
STOCKS_CRYPTO_INTERVAL_SEC = 5 * 60  # 5 minutes

import pandas as pd
import requests
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# CoinGecko symbol to API id mapping (case-insensitive via .get(s.upper()))
_COINGECKO_IDS = {
    "BTC": "bitcoin",
    "ETH": "ethereum",
    "XRP": "ripple",
    "SOL": "solana",
    "ADA": "cardano",
    "XLM": "stellar",
    "XTZ": "tezos",
    "DOGE": "dogecoin",
    "CBETH": "coinbase-wrapped-staked-eth",
    "VARA": "varas",
    "USDC": "usd-coin",
    "DAI": "dai",
    "VET": "vechain",
    "GRT": "the-graph",
    "RLY": "rally-2",
    "CLV": "clover-finance",
    "SKL": "skale",
    "MLN": "melon",
    "AMP": "amp-token",
}
COINGECKO_IDS = _COINGECKO_IDS


def load_config(config_path: Path) -> dict:
    """Load configuration from JSON."""
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_effective_api_keys(config: dict) -> dict:
    """
    Return API keys with environment overrides. Env vars take precedence over config
    so secrets can live in env (e.g. COINBASE_KEY_NAME, COINBASE_PRIVATE_KEY, GOLDAPI_IO)
    and config.json can stay out of version control or cloud sync.
    """
    keys = dict(config.get("api_keys") or {})
    if os.environ.get("GOLDAPI_IO"):
        keys["goldapi_io"] = os.environ.get("GOLDAPI_IO", "")
    if os.environ.get("COINBASE_KEY_NAME"):
        keys["coinbase_key_name"] = os.environ.get("COINBASE_KEY_NAME", "")
    if os.environ.get("COINBASE_PRIVATE_KEY"):
        keys["coinbase_private_key"] = os.environ.get("COINBASE_PRIVATE_KEY", "")
    return keys


def is_market_hours():
    """True if US equity market is open (9:30 AM - 4:00 PM ET, Mon-Fri)."""
    try:
        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo("America/New_York"))
    except ImportError:
        now = datetime.now()
    if now.weekday() >= 5:
        return False
    return (now.hour, now.minute) >= (9, 30) and (now.hour, now.minute) <= (16, 0)


def fetch_stock_prices(tickers):
    """Fetch current prices for stocks/ETFs via yfinance."""
    prices = {}
    if not tickers:
        return prices
    # Batch download first
    try:
        data = yf.download(tickers, group_by="ticker", auto_adjust=True, progress=False, threads=False)
        if data is not None and not data.empty:
            if len(tickers) == 1:
                # Single ticker: data is a simple DataFrame
                if "Close" in data.columns:
                    close = data["Close"].dropna()
                    if len(close) > 0:
                        prices[tickers[0]] = float(close.iloc[-1])
            elif isinstance(data.columns, pd.MultiIndex):
                # Multi-ticker: MultiIndex columns (ticker, field)
                for t in tickers:
                    try:
                        close = data[t]["Close"].dropna() if t in data.columns.get_level_values(0) else None
                        if close is not None and len(close) > 0:
                            prices[t] = float(close.iloc[-1])
                    except (KeyError, TypeError):
                        pass
            # If batch returned flat columns, skip — fallback below will handle it
    except Exception:
        pass
    # Individual fallback for any missing tickers
    for t in tickers:
        if t not in prices:
            try:
                ticker = yf.Ticker(t)
                hist = ticker.history(period="5d")
                if hist is not None and len(hist) > 0:
                    prices[t] = float(hist["Close"].iloc[-1])
            except Exception:
                pass
    return prices


def fetch_crypto_prices(symbols, retries: int = 2):
    """Fetch crypto prices from CoinGecko (free, no key). Retries on empty/429."""
    prices = {}
    if not symbols:
        return prices
    ids = [COINGECKO_IDS.get((s or "").upper(), (s or "").lower()) for s in symbols]
    ids_param = ",".join(ids)
    url = f"https://api.coingecko.com/api/v3/simple/price?ids={ids_param}&vs_currencies=usd"
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=15)
            if r.status_code == 200:
                data = r.json()
                for sym, cg_id in zip(symbols, ids):
                    if cg_id in data and "usd" in data[cg_id]:
                        prices[sym] = float(data[cg_id]["usd"])
                if prices or attempt == retries - 1:
                    break
            if r.status_code == 429 or (r.status_code != 200 and attempt < retries - 1):
                time.sleep(2)
        except Exception:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                pass
    return prices


def fetch_coinbase_balances(
    api_key_name: str,
    private_key_pem: str,
    verbose: bool = False,
) -> Optional[list[dict]]:
    """
    Fetch crypto balances from Coinbase Advanced Trade API.
    Returns list of {"symbol": "BTC", "qty": 0.1} or None if not configured/fails.
    """
    if not (api_key_name and api_key_name.strip() and private_key_pem and private_key_pem.strip()):
        return None
    key_pem = private_key_pem.strip()
    # Ensure PEM newlines: .env may store literal \n; SDK needs real newlines
    if "\\n" in key_pem:
        key_pem = key_pem.replace("\\n", "\n")
    try:
        from coinbase.rest import RESTClient

        client = RESTClient(api_key=api_key_name.strip(), api_secret=key_pem)
        out = []
        cursor = None
        while True:
            resp = client.get_accounts(limit=250, cursor=cursor)
            # Handle dict or object response
            if hasattr(resp, "get"):
                accounts = resp.get("accounts", [])
                cursor = resp.get("cursor")
                has_next = resp.get("has_next", False)
            else:
                accounts = getattr(resp, "accounts", []) or []
                cursor = getattr(resp, "cursor", None)
                has_next = getattr(resp, "has_next", False)
            for acc in accounts:
                if hasattr(acc, "get"):
                    currency = acc.get("currency", "")
                    bal = acc.get("available_balance") or acc.get("balance") or {}
                    val = bal.get("value") if isinstance(bal, dict) else getattr(bal, "value", None)
                else:
                    currency = getattr(acc, "currency", "")
                    bal = getattr(acc, "available_balance", None) or getattr(acc, "balance", None)
                    val = bal.get("value") if isinstance(bal, dict) else (getattr(bal, "value", None) if bal else None)
                if not currency or currency.upper() == "USD":
                    continue
                try:
                    qty = float(val) if val is not None else 0.0
                except (TypeError, ValueError):
                    qty = 0.0
                if qty > 0:
                    out.append({"symbol": currency.upper(), "qty": qty})
            if not has_next or not cursor:
                break
        if verbose and out:
            print(f"  Coinbase: {len(out)} balances")
        return out if out else None
    except Exception as e:
        if verbose:
            print(f"  Coinbase: skip ({e})")
        return None


def fetch_treasury_yields(verbose: bool = False) -> dict:
    """Fetch 10-year and 2-year Treasury yields via yfinance. Returns {tnx_10y: float, irx_2y: float}."""
    out = {}
    # ^TNX = 10-year, 2YY=F = 2-year yield futures (tracks 2Y closely)
    symbols = [("^TNX", "tnx_10y"), ("2YY=F", "tnx_2y")]
    for sym, key in symbols:
        try:
            ticker = yf.Ticker(sym)
            hist = ticker.history(period="5d")
            if hist is not None and len(hist) > 0:
                out[key] = float(hist["Close"].iloc[-1])
        except Exception:
            pass
    return out


# Reasonable price ranges for sanity checking (updated periodically)
_METALS_SANITY = {
    "GOLD":   (1500, 15000),   # Gold per oz — reject if outside this range
    "SILVER": (15, 300),       # Silver per oz
}


def _metals_sane(prices: dict) -> dict:
    """Remove metals prices that fall outside plausible ranges."""
    clean = {}
    for key, val in prices.items():
        lo, hi = _METALS_SANITY.get(key, (0, 1e9))
        if val is not None and lo <= val <= hi:
            clean[key] = val
    return clean


def fetch_metals_prices(gold_api_key: Optional[str], verbose: bool = False) -> dict:
    """Fetch gold and silver spot prices. Try GoldAPI.io first, then yfinance futures as fallback.
    Returns only prices that pass sanity checks."""
    prices = {}
    # Try GoldAPI.io first (if key provided and quota available)
    if gold_api_key and gold_api_key.strip():
        headers = {"x-access-token": gold_api_key.strip(), "Content-Type": "application/json"}
        for metal in ["XAU", "XAG"]:
            try:
                url = f"https://www.goldapi.io/api/{metal}/USD"
                r = requests.get(url, headers=headers, timeout=10)
                if r.status_code == 200:
                    data = r.json()
                    price = data.get("price")
                    if price is not None:
                        prices["GOLD" if metal == "XAU" else "SILVER"] = float(price)
                elif verbose:
                    print(f"  GoldAPI {metal}: {r.status_code}")
            except Exception as e:
                if verbose:
                    print(f"  GoldAPI {metal}: {e}")
    # Fallback: yfinance gold/silver futures (free, no key)
    if "GOLD" not in prices or "SILVER" not in prices:
        try:
            for sym, key in [("GC=F", "GOLD"), ("SI=F", "SILVER")]:
                if key in prices:
                    continue
                ticker = yf.Ticker(sym)
                hist = ticker.history(period="5d")
                if hist is not None and len(hist) > 0:
                    prices[key] = float(hist["Close"].iloc[-1])
            if verbose and prices:
                print(f"  Metals: yfinance fallback")
        except Exception as e:
            if verbose:
                print(f"  Metals yfinance: {e}")
    # Reject implausible values (e.g. gold=2500 when it should be ~5000)
    clean = _metals_sane(prices)
    if verbose and len(clean) < len(prices):
        dropped = {k: v for k, v in prices.items() if k not in clean}
        print(f"  Metals: dropped implausible values {dropped}")
    return clean


def compute_holdings_values(
    config: dict,
    stock_prices: dict[str, float],
    crypto_prices: dict[str, float],
    metals_prices: dict[str, float],
) -> tuple[list[dict], float]:
    """Compute current value of all holdings. Returns list of holding rows and total portfolio value."""
    holdings = []
    total = 0.0

    # Fidelity holdings (stocks/ETFs)
    for h in config.get("holdings", []):
        ticker = h.get("ticker", "")
        asset_class = h.get("asset_class", "Equities")
        val = h.get("value_override")
        qty = h.get("qty")
        # Prefer live price * qty when available (real-time valuation)
        if ticker in stock_prices and qty is not None and float(qty) > 0:
            value = stock_prices[ticker] * float(qty)
        elif val is not None:
            # Fallback to value_override (SPAXX cash, positions without live price)
            value = float(val)
        else:
            value = 0.0
        if value > 0:
            holdings.append({
                "account": h.get("account", "Fidelity"),
                "ticker": ticker,
                "asset_class": asset_class,
                "qty": qty,
                "value": value,
                "notes": h.get("notes", ""),
            })
            total += value

    # Blended accounts (use static values - user updates periodically)
    for b in config.get("blended_accounts", []):
        val = float(b.get("value", 0))
        holdings.append({
            "account": b.get("name", ""),
            "ticker": "-",
            "asset_class": b.get("asset_class", "ManagedBlend"),
            "qty": None,
            "value": val,
            "notes": "Manual update",
        })
        total += val

    # Crypto
    gold_price = metals_prices.get("GOLD") or 5000  # fallback
    silver_price = metals_prices.get("SILVER") or 80  # fallback

    for c in config.get("crypto_holdings", []):
        sym = c.get("symbol", "")
        qty = float(c.get("qty", 0))
        price = crypto_prices.get(sym, 0)
        value = qty * price
        holdings.append({
            "account": "Crypto",
            "ticker": sym,
            "asset_class": "Crypto",
            "qty": qty,
            "value": value,
            "notes": "",
        })
        total += value

    # Physical metals — read from physical_metals array, fallback to legacy inputs fields
    phys_metals = config.get("physical_metals", [])
    if phys_metals:
        gold_oz = sum(float(m.get("qty_oz", 0)) for m in phys_metals if m.get("metal", "").lower() == "gold")
        silver_oz = sum(float(m.get("qty_oz", 0)) for m in phys_metals if m.get("metal", "").lower() == "silver")
    else:
        gold_oz = config.get("inputs", {}).get("physical_gold_oz", 0)
        silver_oz = config.get("inputs", {}).get("physical_silver_oz", 0)

    phys_gold_val = gold_oz * gold_price
    phys_silver_val = silver_oz * silver_price
    holdings.append({
        "account": "Physical",
        "ticker": "PHYS_GOLD",
        "asset_class": "Gold",
        "qty": gold_oz,
        "value": phys_gold_val,
        "notes": f"Spot ~${gold_price:.0f}/oz",
    })
    holdings.append({
        "account": "Physical",
        "ticker": "PHYS_SILVER",
        "asset_class": "Silver",
        "qty": silver_oz,
        "value": phys_silver_val,
        "notes": f"Spot ~${silver_price:.0f}/oz",
    })
    total += phys_gold_val + phys_silver_val

    return holdings, total


def aggregate_by_bucket(holdings: list[dict]) -> dict[str, float]:
    """Map asset classes to target buckets and sum values."""
    mapping = {
        "Cash": "Cash",
        "Equities": "Equities",
        "Gold": "Gold",
        "Silver": "Silver",
        "Crypto": "Crypto",
        "RealEstate": "RealAssets",
        "Art": "Art",
        "ManagedBlend": "Equities",
        "RetirementBlend": "Equities",
        "RealAssets": "RealAssets",
    }
    buckets = {"Cash": 0, "Equities": 0, "Gold": 0, "Silver": 0, "Crypto": 0, "RealAssets": 0, "Art": 0}
    for h in holdings:
        ac = h.get("asset_class", "Equities")
        bucket = mapping.get(ac, "Equities")
        buckets[bucket] = buckets.get(bucket, 0) + h.get("value", 0)
    return buckets


def get_next_buys(config: dict, period_index: int) -> dict:
    """Get the dollar amounts for next bi-weekly contribution."""
    plan = config.get("contribution_plan", {})
    tactical_weeks = plan.get("tactical_weeks", 12)
    periods_tactical = tactical_weeks // 2
    is_tactical = period_index < periods_tactical
    if is_tactical:
        split = plan.get("tactical_split", {})
        return {
            "gold_etf": split.get("gold_etf", 300),
            "gold_phys_save": split.get("gold_phys_save", 150),
            "silver_etf": split.get("silver_etf", 180),
            "silver_phys_save": split.get("silver_phys_save", 120),
            "crypto": split.get("crypto", 360),
            "equities": split.get("equities", 390),
            "real_assets": split.get("real_assets", 180),
            "cash": split.get("cash", 300),
        }
    else:
        split = plan.get("catchup_split", {})
        return {
            "gold_etf": split.get("gold_etf", 550),
            "gold_phys_save": 0,
            "silver_etf": split.get("silver", 250) // 2,
            "silver_phys_save": split.get("silver", 250) // 2,
            "crypto": split.get("crypto", 250),
            "equities": split.get("equities", 250),
            "real_assets": split.get("real_assets", 250),
            "cash": split.get("cash", 250),
        }


def generate_dashboard_html(
    config: dict,
    holdings: list[dict],
    buckets: dict[str, float],
    total: float,
    metals_prices: dict[str, float],
) -> str:
    """Generate the HTML dashboard."""
    targets = config.get("targets", {}).get("tactical", {})
    next_buys = get_next_buys(config, 0)
    gold_price = metals_prices.get("GOLD") or 5000
    silver_price = metals_prices.get("SILVER") or 80

    alloc_rows = ""
    for bucket, value in buckets.items():
        if total > 0:
            pct = 100 * value / total
        else:
            pct = 0
        tgt = targets.get(bucket, {}).get("target", 0)
        drift = pct - tgt
        drift_class = "over" if drift > 5 else ("under" if drift < -5 else "ok")
        alloc_rows += f"""
        <tr>
            <td>{bucket}</td>
            <td>${value:,.0f}</td>
            <td>{pct:.1f}%</td>
            <td>{tgt}%</td>
            <td class="{drift_class}">{drift:+.1f}%</td>
        </tr>"""

    links_html = ""
    for label, urls in config.get("links", {}).items():
        if urls:
            links_html += f'<p><strong>{label.replace("_", " ").title()}:</strong> '
            links_html += ", ".join(f'<a href="{u}" target="_blank" rel="noopener">Link</a>' for u in urls[:3])
            links_html += "</p>"

    # Budget section
    budget = config.get("budget", {})
    monthly_income = budget.get("monthly_income", 0)
    categories = budget.get("categories", [])
    budget_rows = ""
    total_limit = 0
    for cat in categories:
        limit = cat.get("limit", 0)
        total_limit += limit
        pct_used = "" if limit == 0 else "—"  # User enters actual in Excel
        budget_rows += f"<tr><td>{cat.get('name', '')}</td><td>${limit:,.0f}</td><td>{pct_used}</td></tr>"
    budget_html = f"""
    <div class="card">
        <h2>Budget</h2>
        <p>Monthly income: ${monthly_income:,.0f} &nbsp;|&nbsp; Target savings: $4,300/bi-weekly</p>
        <table>
            <thead><tr><th>Category</th><th>Limit</th><th>% Used</th></tr></thead>
            <tbody>{budget_rows}</tbody>
        </table>
        <p style="font-size: 0.9rem; color: #8b949e;">Enter actual spending in Excel Budget tab.</p>
    </div>""" if categories else ""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Nickel&amp;Dime Dashboard</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', system-ui, sans-serif; max-width: 900px; margin: 0 auto; padding: 24px; background: #0f1419; color: #e6edf3; }}
        h1 {{ color: #58a6ff; }}
        h2 {{ color: #8b949e; font-size: 1rem; text-transform: uppercase; margin-top: 32px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
        th, td {{ padding: 10px 12px; text-align: left; border-bottom: 1px solid #30363d; }}
        th {{ color: #8b949e; font-weight: 600; }}
        .over {{ color: #f85149; }}
        .under {{ color: #3fb950; }}
        .ok {{ color: #8b949e; }}
        .total {{ font-size: 1.5rem; font-weight: 700; color: #58a6ff; }}
        .card {{ background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 20px; margin-bottom: 16px; }}
        .buys {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap: 12px; }}
        .buy-item {{ background: #21262d; padding: 12px; border-radius: 6px; }}
        .buy-item strong {{ color: #58a6ff; }}
        a {{ color: #58a6ff; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .meta {{ color: #8b949e; font-size: 0.9rem; margin-top: 24px; }}
    </style>
</head>
<body>
    <h1>Nickel&amp;Dime Dashboard</h1>
    <p class="meta">Last updated: {datetime.now().strftime("%Y-%m-%d %H:%M")} &nbsp;|&nbsp;
        <button id="refresh-btn" onclick="refreshPrices()" style="cursor:pointer;padding:6px 12px;background:#238636;color:#fff;border:none;border-radius:6px;font-size:0.9rem;">Refresh prices</button>
        <a href="/balances" style="margin-left:12px;color:#58a6ff;">Update balances</a>
    </p>

    <div class="card">
        <h2>Portfolio Summary</h2>
        <p class="total">Total: ${total:,.2f}</p>
        <p>Gold spot: ${gold_price:,.0f}/oz &nbsp;|&nbsp; Silver spot: ${silver_price:,.1f}/oz</p>
    </div>

    <div class="card">
        <h2>Allocation vs Target</h2>
        <table>
            <thead><tr><th>Bucket</th><th>Value</th><th>Current %</th><th>Target %</th><th>Drift</th></tr></thead>
            <tbody>{alloc_rows}</tbody>
        </table>
    </div>

    <div class="card">
        <h2>Next Bi-Weekly Buys ($2,000)</h2>
        <div class="buys">
            <div class="buy-item"><strong>Gold ETF</strong><br>${next_buys['gold_etf']}</div>
            <div class="buy-item"><strong>Gold Phys Save</strong><br>${next_buys['gold_phys_save']}</div>
            <div class="buy-item"><strong>Silver ETF</strong><br>${next_buys['silver_etf']}</div>
            <div class="buy-item"><strong>Silver Phys Save</strong><br>${next_buys['silver_phys_save']}</div>
            <div class="buy-item"><strong>Crypto</strong><br>${next_buys['crypto']} (70% BTC / 30% ETH)</div>
            <div class="buy-item"><strong>Equities</strong><br>${next_buys['equities']}</div>
            <div class="buy-item"><strong>Real Assets</strong><br>${next_buys['real_assets']}</div>
            <div class="buy-item"><strong>Cash</strong><br>${next_buys['cash']}</div>
        </div>
        <p style="margin-top: 12px; font-size: 0.9rem;">Crypto ladder: 50% now, 30% at -10%, 20% at -20%</p>
    </div>

    {budget_html}

    <div class="card">
        <h2>Quick Links</h2>
        {links_html}
    </div>

    <div class="card">
        <h2>Rules</h2>
        <ul style="line-height: 1.8; color: #8b949e;">
            <li>Stop Stash/Acorns contributions. Route all new $ to Fidelity.</li>
            <li>No new SLV/GLD; use PSLV and GLDM/IAU for new buys.</li>
            <li>Crypto cap: 15% until 2 of 3 signals (real yields falling, QT ends, vol compresses).</li>
            <li>Physical gold threshold: $500; silver: $250 or 5–10 oz batch.</li>
        </ul>
    </div>
    <script>
        function refreshPrices() {{
            var btn = document.getElementById('refresh-btn');
            btn.disabled = true;
            btn.textContent = 'Updating...';
            fetch('/refresh', {{ method: 'POST' }})
                .then(function(r) {{ if (r.ok) location.reload(); else throw new Error(r.status); }})
                .catch(function() {{
                    btn.textContent = 'Refresh prices (run server first)';
                    btn.disabled = false;
                }});
        }}
    </script>
</body>
</html>"""
    return html


def update_excel(
    wb_path: Path,
    config: dict,
    holdings: list[dict],
    buckets: dict,
    total: float,
):
    """Create or update the Excel workbook."""
    if wb_path.exists():
        wb = load_workbook(wb_path)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)

    # Ensure sheets exist (History is used for dashboard-save audit log)
    sheet_names = [s.title for s in wb.worksheets]
    for name in ["History", "Holdings", "Dashboard", "Inputs", "Targets", "Budget", "ContributionPlan", "Rules"]:
        if name not in sheet_names:
            wb.create_sheet(name, 0 if name == "History" else None)

    # Holdings
    ws = wb["Holdings"]
    ws.delete_rows(1, ws.max_row)
    headers = ["Account", "Ticker", "Asset Class", "Qty", "Value", "Notes"]
    ws.append(headers)
    for h in holdings:
        ws.append([
            h.get("account", ""),
            h.get("ticker", ""),
            h.get("asset_class", ""),
            h.get("qty"),
            round(h.get("value", 0), 2),
            h.get("notes", ""),
        ])

    # Dashboard summary
    ws = wb["Dashboard"]
    ws.delete_rows(1, ws.max_row)
    ws.append(["Portfolio Total", total])
    ws.append([])
    ws.append(["Bucket", "Value", "Current %", "Target %", "Drift"])
    targets = config.get("targets", {}).get("tactical", {})
    for bucket, value in buckets.items():
        pct = 100 * value / total if total > 0 else 0
        tgt = targets.get(bucket, {}).get("target", 0)
        ws.append([bucket, value, round(pct, 1), tgt, round(pct - tgt, 1)])

    # Inputs
    ws = wb["Inputs"]
    ws.delete_rows(1, ws.max_row)
    inp = config.get("inputs", {})
    contrib = config.get("contribution", {})
    ws.append(["Setting", "Value"])
    phys_metals_exp = config.get("physical_metals", [])
    gold_oz_exp = sum(float(m.get("qty_oz", 0)) for m in phys_metals_exp if m.get("metal", "").lower() == "gold")
    silver_oz_exp = sum(float(m.get("qty_oz", 0)) for m in phys_metals_exp if m.get("metal", "").lower() == "silver")
    ws.append(["Physical Gold (oz)", gold_oz_exp or inp.get("physical_gold_oz", 0.2)])
    ws.append(["Physical Silver (oz)", silver_oz_exp or inp.get("physical_silver_oz", 15)])
    ws.append(["Crypto Cap %", inp.get("crypto_cap_percent", 15)])
    ws.append(["Contribution Amount", contrib.get("amount", 2000)])
    ws.append(["IRA %", contrib.get("ira_percent", 10)])
    ws.append(["Next Contribution Date", inp.get("next_contribution_date", "")])

    # Targets
    ws = wb["Targets"]
    ws.delete_rows(1, ws.max_row)
    ws.append(["Bucket", "Tactical Target %", "Tactical Min", "Tactical Max", "Catchup Target %", "Catchup Min", "Catchup Max"])
    for bucket_name, t in config.get("targets", {}).get("tactical", {}).items():
        c = config.get("targets", {}).get("catchup", {}).get(bucket_name, {})
        ws.append([
            bucket_name,
            t.get("target", 0),
            t.get("min", 0),
            t.get("max", 0),
            c.get("target", 0),
            c.get("min", 0),
            c.get("max", 0),
        ])

    # Budget
    ws = wb["Budget"]
    if ws.max_row <= 1:
        ws.append(["Category", "Monthly Limit", "Actual (enter)", "% Used"])
        for cat in config.get("budget", {}).get("categories", []):
            ws.append([cat.get("name", ""), cat.get("limit", 0), "", ""])
        ws.append(["Monthly Income", config.get("budget", {}).get("monthly_income", 0), "", ""])

    # ContributionPlan - 26 periods (1 year biweekly)
    ws = wb["ContributionPlan"]
    ws.delete_rows(1, ws.max_row)
    plan = config.get("contribution_plan", {})
    start = config.get("inputs", {}).get("next_contribution_date", "2026-02-15")
    try:
        d = datetime.strptime(start, "%Y-%m-%d")
    except Exception:
        d = datetime.now()
    headers = ["Period", "Date", "Phase", "Gold ETF", "Gold Phys Save", "Silver ETF", "Silver Phys Save",
               "Crypto", "Equities", "Real Assets", "Cash", "Physical Buy?"]
    ws.append(headers)
    for i in range(26):
        buys = get_next_buys(config, i)
        dt = d + timedelta(weeks=2 * i) if i > 0 else d
        phase = "Tactical" if i < 6 else "Catch-up"
        ws.append([
            i + 1,
            dt.strftime("%Y-%m-%d"),
            phase,
            buys["gold_etf"],
            buys["gold_phys_save"],
            buys["silver_etf"],
            buys["silver_phys_save"],
            buys["crypto"],
            buys["equities"],
            buys["real_assets"],
            buys["cash"],
            "Check envelope thresholds",
        ])

    # Rules
    ws = wb["Rules"]
    ws.delete_rows(1, ws.max_row)
    rules = [
        "Stop Stash/Acorns contributions. Route all new $ to Fidelity.",
        "No new SLV/GLD; use PSLV and GLDM/IAU for new buys.",
        "Crypto cap: 15% until 2 of 3 signals: real yields falling, QT ends, vol compresses.",
        "Physical gold threshold: $500; silver: $250 or 5-10 oz batch.",
        "Crypto ladder: 50% now, 30% at -10%, 20% at -20%. 70% BTC / 30% ETH.",
        "No new money to XRP/SOL/other alts. Consolidate into BTC/ETH on strength.",
    ]
    ws.append(["Rule"])
    for r in rules:
        ws.append([r])

    wb.save(wb_path)


PRICE_HISTORY_MAX = 3650  # ~10 years of daily data


def _price_cache_path(base: Path) -> Path:
    return base / "price_cache.json"


def load_price_cache(base: Path) -> dict:
    """Load cached prices (metals, crypto). Used when API fails."""
    path = _price_cache_path(base)
    if not path.exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_price_cache(base: Path, metals: dict = None, crypto: dict = None,
                     stocks: dict = None, treasury: dict = None) -> None:
    """Save prices to cache. Merges with existing cache."""
    path = _price_cache_path(base)
    cache = load_price_cache(base)
    if metals:
        cache["metals"] = metals
        cache["metals_updated"] = datetime.now().isoformat()
    if crypto:
        cache["crypto"] = crypto
        cache["crypto_updated"] = datetime.now().isoformat()
    if stocks:
        cache["stocks"] = stocks
        cache["stocks_updated"] = datetime.now().isoformat()
    if treasury:
        cache["treasury"] = treasury
        cache["treasury_updated"] = datetime.now().isoformat()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2)
    except Exception:
        pass


import threading
_price_history_lock = threading.Lock()


def _price_history_path(base: Path) -> Path:
    return base / "price_history.json"


def _read_price_history_unlocked(base: Path) -> list:
    """Read price history from JSON file. Caller must hold _price_history_lock.
    Falls back to .bak file if main file is corrupt, then tries Excel restore."""
    path = _price_history_path(base)
    bak_path = Path(str(path) + ".bak")
    history = []
    for p in (path, bak_path):
        if p.exists():
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                loaded = data.get("history", []) if isinstance(data, dict) else (data if isinstance(data, list) else [])
                if len(loaded) > len(history):
                    history = loaded
            except Exception:
                continue
    if len(history) <= 1:
        if restore_price_history_from_excel(base):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                restored = data.get("history", []) if isinstance(data, dict) else data
                if len(restored) > len(history):
                    history = restored
                    print(f"[Startup] Restored {len(history)} price history entries from Excel backup")
            except Exception:
                pass
    return history


def load_price_history(base: Path) -> list:
    """Load price history. Thread-safe. Auto-restores from Excel backup if JSON is missing/empty."""
    with _price_history_lock:
        return _read_price_history_unlocked(base)


def append_price_history(
    base: Path,
    total: float,
    metals_prices: dict,
    treasury_yields: dict,
    gold_silver_ratio: Optional[float],
) -> None:
    """Append/update daily snapshot with OHLC tracking. Thread-safe.
    
    Each day gets one entry with open/high/low/close for the portfolio total.
    Multiple refreshes in a day update the high, low, and close values.
    Keeps last PRICE_HISTORY_MAX entries.
    Skips the update if the new total deviates >10% from the last known value
    (likely bad API data).
    """
    with _price_history_lock:
        gold = metals_prices.get("GOLD")
        silver = metals_prices.get("SILVER")
        today = datetime.now().strftime("%Y-%m-%d")
        now_time = datetime.now().strftime("%H:%M")
        rounded_total = round(total, 2)

        history = _read_price_history_unlocked(base)

        if history:
            last_total = history[-1].get("close") or history[-1].get("total") or 0
            if last_total > 0:
                pct_change = abs(rounded_total - last_total) / last_total
                if pct_change > 0.10:
                    return

        if history and history[-1].get("date", "")[:10] == today:
            entry = history[-1]
            entry["close"] = rounded_total
            entry["total"] = rounded_total
            if rounded_total > entry.get("high", rounded_total):
                entry["high"] = rounded_total
            if rounded_total < entry.get("low", rounded_total):
                entry["low"] = rounded_total
            entry["last_update"] = now_time
            entry["gold"] = round(gold, 2) if gold is not None else entry.get("gold")
            entry["silver"] = round(silver, 2) if silver is not None else entry.get("silver")
            entry["gold_silver_ratio"] = gold_silver_ratio if gold_silver_ratio is not None else entry.get("gold_silver_ratio")
            entry["tnx_10y"] = round(treasury_yields.get("tnx_10y"), 2) if treasury_yields.get("tnx_10y") is not None else entry.get("tnx_10y")
            entry["tnx_2y"] = round(treasury_yields.get("tnx_2y"), 2) if treasury_yields.get("tnx_2y") is not None else entry.get("tnx_2y")
        else:
            prev_close = history[-1].get("close") or history[-1].get("total") if history else None
            day_open = prev_close if prev_close and prev_close > 0 else rounded_total
            entry = {
                "date": today,
                "total": rounded_total,
                "open": day_open,
                "high": max(day_open, rounded_total),
                "low": min(day_open, rounded_total),
                "close": rounded_total,
                "last_update": now_time,
                "gold": round(gold, 2) if gold is not None else None,
                "silver": round(silver, 2) if silver is not None else None,
                "gold_silver_ratio": gold_silver_ratio,
                "tnx_10y": round(treasury_yields.get("tnx_10y"), 2) if treasury_yields.get("tnx_10y") is not None else None,
                "tnx_2y": round(treasury_yields.get("tnx_2y"), 2) if treasury_yields.get("tnx_2y") is not None else None,
            }
            history.append(entry)

        history = history[-PRICE_HISTORY_MAX:]
        _write_price_history_safe(base, history)


def _write_price_history_safe(base: Path, history: list) -> None:
    """Write price history using atomic rename with backup. Refuses to overwrite
    a longer history with a shorter one (protection against corruption cascades)."""
    import tempfile, shutil
    path = _price_history_path(base)
    bak_path = Path(str(path) + ".bak")

    # Safety: never replace a longer file with fewer entries
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                existing = json.load(f)
            existing_len = len(existing.get("history", []) if isinstance(existing, dict) else existing)
            if len(history) < existing_len - 1:
                print(f"[History] SAFETY: refusing to overwrite {existing_len} entries with {len(history)}")
                return
        except Exception:
            pass

    # Atomic write: write to temp file in same directory, then rename
    try:
        if path.exists():
            shutil.copy2(path, bak_path)
        fd, tmp_path = tempfile.mkstemp(dir=str(base), suffix=".json.tmp")
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as f:
                json.dump({"history": history}, f, indent=2)
            Path(tmp_path).replace(path)
        except Exception:
            Path(tmp_path).unlink(missing_ok=True)
            raise
    except Exception as e:
        print(f"[History] Write failed: {e}")

    # Only sync to Excel if we have meaningful data
    if len(history) >= 2:
        _sync_price_history_to_excel(base, history)


_PRICE_HISTORY_COLS = ["date", "total", "open", "high", "low", "close",
                       "last_update", "gold", "silver", "gold_silver_ratio",
                       "tnx_10y", "tnx_2y"]


def _sync_price_history_to_excel(base: Path, history: list) -> None:
    """Write full price history to a PriceHistory sheet in the Excel workbook."""
    wb_path = base / "Curtis_WealthOS.xlsx"
    try:
        if wb_path.exists():
            wb = load_workbook(wb_path)
        else:
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)
        if "PriceHistory" not in [s.title for s in wb.worksheets]:
            wb.create_sheet("PriceHistory")
        ws = wb["PriceHistory"]
        ws.delete_rows(1, ws.max_row)
        ws.append(_PRICE_HISTORY_COLS)
        for entry in history:
            ws.append([entry.get(c) for c in _PRICE_HISTORY_COLS])
        wb.save(wb_path)
    except Exception:
        pass


def restore_price_history_from_excel(base: Path) -> bool:
    """Rebuild price_history.json from the PriceHistory sheet in Excel.
    Uses zipfile+XML to bypass corrupted sheets that crash openpyxl."""
    import zipfile
    import xml.etree.ElementTree as ET
    wb_path = base / "Curtis_WealthOS.xlsx"
    if not wb_path.exists():
        return False
    try:
        z = zipfile.ZipFile(str(wb_path))
        # Find PriceHistory sheet by reading workbook.xml for sheet->rId mapping
        wb_xml = z.open("xl/workbook.xml").read()
        root = ET.fromstring(wb_xml)
        xns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        target_rid = None
        for s in root.findall(f".//{{{xns}}}sheet"):
            if s.get("name") == "PriceHistory":
                target_rid = s.get(f"{{{rns}}}id")
                break
        if not target_rid:
            z.close()
            return False
        # Resolve rId to file path
        rels_xml = z.open("xl/_rels/workbook.xml.rels").read()
        rels_root = ET.fromstring(rels_xml)
        sheet_file = None
        for rel in rels_root:
            if rel.get("Id") == target_rid:
                sheet_file = rel.get("Target").lstrip("/")
                if not sheet_file.startswith("xl/"):
                    sheet_file = "xl/" + sheet_file
                break
        if not sheet_file:
            z.close()
            return False
        tree = ET.parse(z.open(sheet_file))
        ns = {"s": xns}
        rows = tree.getroot().findall(".//s:row", ns)
        cols = _PRICE_HISTORY_COLS

        def col_to_idx(col_ref):
            result = 0
            for ch in col_ref:
                if ch.isalpha():
                    result = result * 26 + (ord(ch.upper()) - ord('A') + 1)
            return result - 1

        history = []
        for row_idx, row in enumerate(rows):
            if row_idx == 0:
                continue
            cells = row.findall("s:c", ns)
            entry = {}
            for c in cells:
                ref = c.get("r", "")
                ci = col_to_idx("".join(ch for ch in ref if ch.isalpha()))
                if ci >= len(cols):
                    continue
                col_name = cols[ci]
                # Handle inline strings
                is_elem = c.find("s:is", ns)
                if is_elem is not None:
                    t_elem = is_elem.find("s:t", ns)
                    val = t_elem.text if t_elem is not None and t_elem.text else None
                else:
                    v_elem = c.find("s:v", ns)
                    val = v_elem.text if v_elem is not None and v_elem.text else None
                if val is None:
                    continue
                if col_name in ("date", "last_update"):
                    entry[col_name] = val
                else:
                    try:
                        entry[col_name] = round(float(val), 2)
                    except ValueError:
                        entry[col_name] = val
            if entry.get("date"):
                history.append(entry)
        z.close()
        if not history:
            return False
        path = _price_history_path(base)
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"history": history}, f, indent=2)
        return True
    except Exception:
        return False


def append_history_log(base: Path, action: str, details: str = "") -> None:
    """Append a row to the History sheet in the Excel workbook for audit trail."""
    wb_path = base / "Curtis_WealthOS.xlsx"
    if not wb_path.exists():
        return
    try:
        wb = load_workbook(wb_path)
        if "History" not in [s.title for s in wb.worksheets]:
            wb.create_sheet("History", 0)
        ws = wb["History"]
        if ws.max_row == 0 or (ws.max_row == 1 and not any(ws.cell(1, c).value for c in range(1, 4))):
            ws.append(["Date", "Action", "Details"])
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), action, details])
        wb.save(wb_path)
    except Exception:
        pass


def get_dashboard_data_cached(base: Path, config: dict) -> dict:
    """
    Fast version: uses only cached/local data, NO network calls.
    Returns instantly for snappy initial page load.
    """
    price_cache = load_price_cache(base)
    metals_prices = price_cache.get("metals", {"GOLD": 5000, "SILVER": 80})
    if not metals_prices:
        metals_prices = {"GOLD": 5000, "SILVER": 80}
    gold = metals_prices.get("GOLD") or 5000
    silver = metals_prices.get("SILVER") or 80
    gold_silver_ratio = round(gold / silver, 2) if silver and silver > 0 else None

    crypto_prices = price_cache.get("crypto", {})
    stock_prices = price_cache.get("stocks", {})

    # Treasury yields from cache
    treasury_yields = price_cache.get("treasury", {})

    bootstrap = {
        "GLD": 4445.93, "SLV": 2180.06, "SPY": 6456.07, "VTI": 4793.17, "TSLA": 3866.80, "DECK": 2393.60,
        "BROS": 2568.62, "COIN": 2107.67, "AMD": 1699.95, "AAPL": 1687.91, "AMZN": 1051.24, "RDDT": 976.88,
        "VALE": 951.68, "FCAEX": 905.59, "NVDA": 864.67, "GOOGL": 762.86, "ARR": 677.81, "STUB": 517.72,
        "F": 499.24, "ORCL": 438.60, "ARKX": 1014.17, "ARKF": 281.86, "AZO": 200.03, "COST": 192.75,
        "CCL": 400.19, "UWMC": 254.80, "SCHD": 885.41, "QQQX": 2022.18,
    }
    for h in config.get("holdings", []):
        t = h.get("ticker")
        if t and t != "SPAXX" and h.get("value_override") is None and t in bootstrap:
            h["value_override"] = bootstrap[t]

    holdings, total = compute_holdings_values(config, stock_prices, crypto_prices, metals_prices)
    buckets = aggregate_by_bucket(holdings)
    price_history = load_price_history(base)
    return {
        "holdings": holdings,
        "buckets": buckets,
        "total": total,
        "metals_prices": metals_prices,
        "stock_prices": stock_prices,
        "crypto_prices": crypto_prices,
        "config": config,
        "treasury_yields": treasury_yields,
        "gold_silver_ratio": gold_silver_ratio,
        "price_history": price_history,
    }


def get_dashboard_data(base: Path, config: dict, verbose: bool = False) -> dict:
    """
    Fetch prices, compute holdings and buckets. Syncs Coinbase if keys present.
    Returns dict with holdings, buckets, total, metals_prices, config (possibly updated).
    Does not write HTML or Excel.
    """
    config_path = base / "config.json"
    api_keys = get_effective_api_keys(config)
    gold_key = api_keys.get("goldapi_io", "")
    price_cache = load_price_cache(base)

    # Coinbase: always sync balances if keys present
    cb_name = (api_keys.get("coinbase_key_name") or "").strip()
    cb_key = (api_keys.get("coinbase_private_key") or "").strip()
    if cb_name and cb_key:
        cb_holdings = fetch_coinbase_balances(cb_name, cb_key, verbose=verbose)
        if cb_holdings is not None and len(cb_holdings) > 0:
            config["crypto_holdings"] = cb_holdings
            try:
                with open(config_path, "w", encoding="utf-8") as f:
                    json.dump(config, f, indent=2)
            except Exception:
                pass

    crypto_symbols = [c["symbol"] for c in config.get("crypto_holdings", [])]

    # Metals: fetch live, fallback to cache, then hardcoded
    metals_prices = fetch_metals_prices(gold_key, verbose=verbose)
    if metals_prices:
        save_price_cache(base, metals=metals_prices)
    else:
        metals_prices = price_cache.get("metals", {})
    if not metals_prices:
        metals_prices = {"GOLD": 5000, "SILVER": 80}
    gold = metals_prices.get("GOLD") or 5000
    silver = metals_prices.get("SILVER") or 80
    gold_silver_ratio = round(gold / silver, 2) if silver and silver > 0 else None
    treasury_yields = fetch_treasury_yields(verbose=verbose)
    if treasury_yields:
        save_price_cache(base, treasury=treasury_yields)
    else:
        treasury_yields = price_cache.get("treasury", {})

    tickers = list({h["ticker"] for h in config.get("holdings", []) if h.get("ticker") and h["ticker"] != "SPAXX"})
    # Always include SPY, DXY, VIX, and Oil for pulse bar / market display
    for sym in ("SPY", "DX-Y.NYB", "^VIX", "CL=F", "HG=F"):
        if sym not in tickers:
            tickers.append(sym)
    # Fetch stock prices (even outside market hours to get latest close)
    stock_prices = fetch_stock_prices(tickers)
    if stock_prices:
        save_price_cache(base, stocks=stock_prices)
    else:
        stock_prices = price_cache.get("stocks", {})

    # Crypto prices: fetch live, fallback to cache
    crypto_prices = fetch_crypto_prices(crypto_symbols)
    if crypto_prices:
        save_price_cache(base, crypto=crypto_prices)
    else:
        crypto_prices = price_cache.get("crypto", {})

    bootstrap = {
        "GLD": 4445.93, "SLV": 2180.06, "SPY": 6456.07, "VTI": 4793.17, "TSLA": 3866.80, "DECK": 2393.60,
        "BROS": 2568.62, "COIN": 2107.67, "AMD": 1699.95, "AAPL": 1687.91, "AMZN": 1051.24, "RDDT": 976.88,
        "VALE": 951.68, "FCAEX": 905.59, "NVDA": 864.67, "GOOGL": 762.86, "ARR": 677.81, "STUB": 517.72,
        "F": 499.24, "ORCL": 438.60, "ARKX": 1014.17, "ARKF": 281.86, "AZO": 200.03, "COST": 192.75,
        "CCL": 400.19, "UWMC": 254.80, "SCHD": 885.41, "QQQX": 2022.18,
    }
    for h in config.get("holdings", []):
        t = h.get("ticker")
        if t and t != "SPAXX" and h.get("value_override") is None and t in bootstrap:
            h["value_override"] = bootstrap[t]

    holdings, total = compute_holdings_values(config, stock_prices, crypto_prices, metals_prices)
    buckets = aggregate_by_bucket(holdings)
    price_history = load_price_history(base)
    return {
        "holdings": holdings,
        "buckets": buckets,
        "total": total,
        "metals_prices": metals_prices,
        "stock_prices": stock_prices,
        "crypto_prices": crypto_prices,
        "config": config,
        "treasury_yields": treasury_yields,
        "gold_silver_ratio": gold_silver_ratio,
        "price_history": price_history,
    }


def run_update(base, config, tickers, crypto_symbols, gold_key, metals_prices=None, fetch_metals=True, verbose=True):
    """Run one update cycle. Returns (total, metals_prices).
    Always syncs Coinbase balances and fetches live prices. Uses cache when APIs fail.
    """
    wb_path = base / "Curtis_WealthOS.xlsx"
    html_path = base / "dashboard.html"
    config_path = base / "config.json"
    price_cache = load_price_cache(base)

    # Always sync Coinbase if keys present
    api_keys = get_effective_api_keys(config)
    cb_name = (api_keys.get("coinbase_key_name") or "").strip()
    cb_key = (api_keys.get("coinbase_private_key") or "").strip()
    if cb_name and cb_key:
        cb_holdings = fetch_coinbase_balances(cb_name, cb_key, verbose=verbose)
        if cb_holdings is not None and len(cb_holdings) > 0:
            config["crypto_holdings"] = cb_holdings
            crypto_symbols[:] = [c["symbol"] for c in cb_holdings]
            try:
                with open(config_path, "w", encoding="utf-8") as f:
                    json.dump(config, f, indent=2)
                if verbose:
                    print("  Coinbase: balances saved to config")
            except Exception as e:
                if verbose:
                    print(f"  Coinbase: could not save config ({e})")

    # Metals: fetch live, fallback to cache
    if fetch_metals:
        metals_prices = fetch_metals_prices(gold_key, verbose=verbose)
        if metals_prices:
            save_price_cache(base, metals=metals_prices)
        else:
            metals_prices = price_cache.get("metals", {})
        if not metals_prices:
            metals_prices = {"GOLD": 5000, "SILVER": 80}
        if verbose:
            status = "live" if metals_prices.get("GOLD", 0) > 0 else "fallback"
            print(f"  Metals: {status}")
    if metals_prices is None:
        metals_prices = {"GOLD": 5000, "SILVER": 80}
    gold = metals_prices.get("GOLD") or 5000
    silver = metals_prices.get("SILVER") or 80
    gold_silver_ratio = round(gold / silver, 2) if silver and silver > 0 else None
    treasury_yields = fetch_treasury_yields(verbose=verbose)
    if treasury_yields:
        save_price_cache(base, treasury=treasury_yields)
    else:
        treasury_yields = price_cache.get("treasury", {})

    # Always include SPY, DXY, VIX, and Oil for pulse bar / market display
    for sym in ("SPY", "DX-Y.NYB", "^VIX", "CL=F", "HG=F"):
        if sym not in tickers:
            tickers.append(sym)
    # Fetch stock prices (even outside market hours to get latest close)
    stock_prices = fetch_stock_prices(tickers)
    if stock_prices:
        save_price_cache(base, stocks=stock_prices)
        if verbose:
            print(f"  Stocks: {len(stock_prices)} prices")
    else:
        stock_prices = price_cache.get("stocks", {})
        if verbose:
            print(f"  Stocks: using cached prices ({len(stock_prices)})")

    # Crypto prices: fetch live, fallback to cache
    crypto_prices = fetch_crypto_prices(crypto_symbols)
    if crypto_prices:
        save_price_cache(base, crypto=crypto_prices)
        if verbose:
            print(f"  Crypto: {len(crypto_prices)} prices")
    else:
        crypto_prices = price_cache.get("crypto", {})
        if verbose:
            print(f"  Crypto: using cached prices ({len(crypto_prices)})")

    bootstrap = {
        "GLD": 4445.93, "SLV": 2180.06, "SPY": 6456.07, "VTI": 4793.17, "TSLA": 3866.80, "DECK": 2393.60,
        "BROS": 2568.62, "COIN": 2107.67, "AMD": 1699.95, "AAPL": 1687.91, "AMZN": 1051.24, "RDDT": 976.88,
        "VALE": 951.68, "FCAEX": 905.59, "NVDA": 864.67, "GOOGL": 762.86, "ARR": 677.81, "STUB": 517.72,
        "F": 499.24, "ORCL": 438.60, "ARKX": 1014.17, "ARKF": 281.86, "AZO": 200.03, "COST": 192.75,
        "CCL": 400.19, "UWMC": 254.80, "SCHD": 885.41, "QQQX": 2022.18,
    }
    for h in config.get("holdings", []):
        t = h.get("ticker")
        if t and t != "SPAXX" and h.get("value_override") is None and t in bootstrap:
            h["value_override"] = bootstrap[t]

    holdings, total = compute_holdings_values(config, stock_prices, crypto_prices, metals_prices)
    buckets = aggregate_by_bucket(holdings)

    append_price_history(base, total, metals_prices, treasury_yields, gold_silver_ratio)

    # Use dashboard.py's render_dashboard so static file matches server (DXY, Economics tab, etc.)
    price_history = load_price_history(base)
    data = {
        "holdings": holdings,
        "buckets": buckets,
        "total": total,
        "metals_prices": metals_prices,
        "stock_prices": stock_prices,
        "crypto_prices": crypto_prices,
        "config": config,
        "treasury_yields": treasury_yields,
        "gold_silver_ratio": gold_silver_ratio,
        "price_history": price_history,
    }
    from dashboard import render_dashboard
    html = render_dashboard(data)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    excel_ok = True
    try:
        update_excel(wb_path, config, holdings, buckets, total)
    except PermissionError:
        excel_ok = False
        if verbose:
            print("  Excel not updated (close Curtis_WealthOS.xlsx and run again to update it)")
    except Exception as e:
        excel_ok = False
        if verbose:
            print(f"  Excel not updated ({e})")

    if verbose:
        print(f"  Total: ${total:,.2f} | Dashboard updated" + (" | Excel updated" if excel_ok else ""))
    return total, metals_prices


def main():
    parser = argparse.ArgumentParser(description="Nickel&Dime Finance Manager")
    parser.add_argument("--watch", action="store_true",
        help="Run in watch mode: update every 5 min (stocks during market hours), metals 4x/day")
    parser.add_argument("--import-csv", type=Path, metavar="FILE",
        help="Import CSV into config (use with --source fidelity|stash|acorns|fundrise)")
    parser.add_argument("--source", choices=["fidelity", "stash", "acorns", "acorns_invest", "acorns_later", "fundrise"],
        help="Source of the CSV for --import-csv")
    args = parser.parse_args()

    base = Path(__file__).resolve().parent
    config_path = base / "config.json"

    if args.import_csv is not None:
        if not args.source:
            print("Use --source fidelity|stash|acorns|fundrise with --import-csv")
            sys.exit(1)
        from csv_import import import_csv
        updated, msg = import_csv(config_path, args.import_csv, args.source)
        print(msg)
        if updated:
            print("Run python finance_manager.py (or refresh-and-open.bat) to refresh the dashboard.")
        sys.exit(0 if updated else 1)

    config = load_config(config_path)

    tickers = list({h["ticker"] for h in config.get("holdings", []) if h.get("ticker") and h["ticker"] != "SPAXX"})
    crypto_symbols = [c["symbol"] for c in config.get("crypto_holdings", [])]
    gold_key = get_effective_api_keys(config).get("goldapi_io", "")

    if not args.watch:
        print("Fetching prices...")
        run_update(base, config, tickers, crypto_symbols, gold_key, fetch_metals=True)
        print("\nDone. Open dashboard.html in your browser.")
        return

    print("Watch mode: stocks every 5 min (market hours), crypto every 5 min, metals ~4x/day. Ctrl+C to stop.")
    metals_prices = None
    last_metals = datetime.min
    try:
        while True:
            now = datetime.now()
            elapsed = (now - last_metals).total_seconds()
            fetch_metals = elapsed >= METALS_INTERVAL_HOURS * 3600

            print(f"\n[{now.strftime('%H:%M:%S')}] Updating...")
            _, metals_prices = run_update(
                base, config, tickers, crypto_symbols, gold_key,
                metals_prices=metals_prices, fetch_metals=fetch_metals, verbose=True
            )
            if fetch_metals:
                last_metals = now

            print(f"  Next update in 5 min")
            time.sleep(STOCKS_CRYPTO_INTERVAL_SEC)
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    main()
